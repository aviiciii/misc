
import openpyxl
import requests
from openpyxl.styles import PatternFill
from tqdm import tqdm

# Load the Excel file
file_path = 'study-jam.xlsx'
wb = openpyxl.load_workbook(file_path)
sheet = wb.active

# Define the column index containing the URLs (assuming column B)
url_column = 11

# Define a red fill for marking cells with 404 errors
red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

# Iterate through rows and check URLs
for row in tqdm(sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=url_column, max_col=url_column)):
    cell = row[0]
    url = cell.value
    
    if url:
        response = requests.head(url)
        if response.status_code != 200:
            cell.fill = red_fill

    

# Save the modified Excel file
output_file_path = 'study-jam-output.xlsx'
wb.save(output_file_path)
print("URL check and cell coloring completed.")